
import matplotlib.pyplot as plt
import openpyxl
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from datetime import datetime

# Get the path to the Excel file
excel_file_path = r"C:\Users\AlphaH\Desktop\Time graph.xlsx"

# Get the current date
current_date = datetime.now().strftime('%Y-%m-%d')

tasks = []
hours_per_task = []

while sum(hours_per_task) < 24:
    task = input("Enter a task (type 'end' to finish): ")
    if task.lower() == 'end':
        break
    
    try:
        hours = float(input(f"Enter hours spent on '{task}': "))
    except ValueError:
        print("Invalid input. Please enter a valid number for hours.")
        continue
    
    tasks.append(task)
    hours_per_task.append(hours)
    
    # Check if the total hours exceed 24
    if sum(hours_per_task) > 24:
        print("Task limit reached. Ending input.")
        break

# Fill the remaining time with "You did nothing"
if sum(hours_per_task) < 24:
    tasks.append("You did nothing")
    hours_per_task.append(24 - sum(hours_per_task))

# Create a pie chart
colors = plt.cm.Pastel1(range(len(tasks)))
plt.pie(hours_per_task, labels=tasks, colors=colors, startangle=90, counterclock=False,
        autopct='%1.f%%', pctdistance=0.85)
plt.axis('equal')
plt.title('Tasks Schedule for 24 Hours')

# Display the ordered time spent on every task
for task, hours in sorted(zip(tasks, hours_per_task), key=lambda x: x[1], reverse=True):
    print(f"{task}: {hours} hours")

# Save the pie chart as an image
chart_image_path = 'pie_chart.png'
plt.savefig(chart_image_path)
plt.close()

# Open the Excel workbook
wb = openpyxl.load_workbook(excel_file_path)

# Create a new worksheet for the chart
ws = wb.create_sheet(title=current_date)

# Add the chart image to the worksheet
img = openpyxl.drawing.image.Image(chart_image_path)
ws.add_image(img, 'A1')

# Insert the date in the Excel file
ws['A3'] = f"Date: {current_date}"

# Save the changes to the Excel file
wb.save(excel_file_path)
wb.close()





